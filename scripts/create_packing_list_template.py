import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
import sys

sys.path.insert(0, str(Path(__file__).parent.parent))

from config.settings import Settings


def create_packing_list_template():
    settings = Settings()
    template_path = settings.get_template_path("packing_list_template.xlsx")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Packing List"
    
    title_font = Font(name='Arial', size=16, bold=True, color='000000')
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    label_font = Font(name='Arial', size=10, bold=True, color='000000')
    data_font = Font(name='Arial', size=10, color='000000')
    
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    total_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = 'PACKING LIST'
    title_cell.font = title_font
    title_cell.alignment = center_alignment
    
    ws['A2'] = 'PO:'
    ws['A2'].font = label_font
    ws['A2'].alignment = left_alignment
    ws['B2'].border = thin_border
    ws['B2'].alignment = center_alignment
    
    ws['A3'] = 'Buyer:'
    ws['A3'].font = label_font
    ws['A3'].alignment = left_alignment
    ws['B3'].border = thin_border
    ws['B3'].alignment = center_alignment
    
    ws['A4'] = 'Ship Date:'
    ws['A4'].font = label_font
    ws['A4'].alignment = left_alignment
    ws['B4'].border = thin_border
    ws['B4'].alignment = center_alignment
    
    ws.row_dimensions[6].height = 5
    
    headers = ['Style', 'Color', 'Size', 'Quantity', 'Carton']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_alignment
    
    for row_idx in range(8, 28):
        for col_idx in range(1, 6):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = center_alignment
            cell.font = data_font
    
    ws['A28'] = 'TOTAL'
    ws['A28'].font = Font(name='Arial', size=11, bold=True, color='000000')
    ws['A28'].fill = total_fill
    ws['A28'].border = thin_border
    ws['A28'].alignment = center_alignment
    
    ws.merge_cells('B28:C28')
    ws['B28'].fill = total_fill
    ws['B28'].border = thin_border
    
    ws['D28'].value = '=SUM(D8:D27)'
    ws['D28'].font = Font(name='Arial', size=11, bold=True, color='000000')
    ws['D28'].fill = total_fill
    ws['D28'].border = thin_border
    ws['D28'].alignment = center_alignment
    ws['D28'].number_format = '#,##0'
    
    ws['E28'].value = '=SUM(E8:E27)'
    ws['E28'].font = Font(name='Arial', size=11, bold=True, color='000000')
    ws['E28'].fill = total_fill
    ws['E28'].border = thin_border
    ws['E28'].alignment = center_alignment
    ws['E28'].number_format = '#,##0'
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[7].height = 20
    ws.row_dimensions[28].height = 20
    
    wb.save(template_path)
    print(f"✅ Đã tạo template tại: {template_path}")


if __name__ == "__main__":
    create_packing_list_template()

