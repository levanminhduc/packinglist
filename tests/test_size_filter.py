import unittest
import tempfile
import openpyxl
from pathlib import Path
import sys

sys.path.insert(0, str(Path(__file__).parent.parent))

from excel_automation.size_filter import SizeFilterManager
from excel_automation.size_filter_config import SizeFilterConfig


class TestSizeFilterConfig(unittest.TestCase):
    
    def setUp(self):
        self.temp_config = tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False)
        self.temp_config.close()
        self.config = SizeFilterConfig(self.temp_config.name)
    
    def tearDown(self):
        Path(self.temp_config.name).unlink(missing_ok=True)
    
    def test_default_config(self):
        self.assertEqual(self.config.get_column(), "F")
        self.assertEqual(self.config.get_start_row(), 19)
        self.assertEqual(self.config.get_end_row(), 59)
        self.assertEqual(self.config.get_sheet_name(), "Sheet1")
    
    def test_update_config(self):
        self.config.update_config("G", 20, 50, "Sheet2")
        self.assertEqual(self.config.get_column(), "G")
        self.assertEqual(self.config.get_start_row(), 20)
        self.assertEqual(self.config.get_end_row(), 50)
        self.assertEqual(self.config.get_sheet_name(), "Sheet2")
    
    def test_validation_start_row_less_than_1(self):
        with self.assertRaises(ValueError):
            self.config.update_config("F", 0, 59, "Sheet1")
    
    def test_validation_start_row_greater_than_end_row(self):
        with self.assertRaises(ValueError):
            self.config.update_config("F", 60, 59, "Sheet1")
    
    def test_validate_config_with_max_row(self):
        self.config.update_config("F", 19, 100, "Sheet1")
        is_valid, msg = self.config.validate_config(max_row=50)
        self.assertFalse(is_valid)
        self.assertIn("vượt quá", msg)
    
    def test_reset_to_defaults(self):
        self.config.update_config("G", 20, 50, "Sheet2")
        self.config.reset_to_defaults()
        self.assertEqual(self.config.get_column(), "F")
        self.assertEqual(self.config.get_start_row(), 19)
        self.assertEqual(self.config.get_end_row(), 59)


class TestSizeFilterManager(unittest.TestCase):
    
    def setUp(self):
        self.temp_file = tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False)
        self.temp_file.close()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        for row in range(1, 19):
            ws[f'F{row}'] = f"Header {row}"
        
        sizes = ["044", "045", "046", "044", "045", "046", "047", "048"]
        for i, size in enumerate(sizes, start=19):
            ws[f'F{i}'] = size
        
        for row in range(27, 60):
            ws[f'F{row}'] = "048"
        
        for row in range(60, 70):
            ws[f'F{row}'] = "Footer"
        
        wb.save(self.temp_file.name)
        wb.close()
        
        self.config = SizeFilterConfig()
    
    def tearDown(self):
        Path(self.temp_file.name).unlink(missing_ok=True)
    
    def test_scan_sizes(self):
        with SizeFilterManager(self.temp_file.name, self.config) as manager:
            sizes = manager.scan_sizes()
            self.assertIn("044", sizes)
            self.assertIn("045", sizes)
            self.assertIn("046", sizes)
            self.assertIn("047", sizes)
            self.assertIn("048", sizes)
    
    def test_get_size_row_mapping(self):
        with SizeFilterManager(self.temp_file.name, self.config) as manager:
            size_rows = manager.get_size_row_mapping()
            self.assertIn("044", size_rows)
            self.assertEqual(len(size_rows["044"]), 2)
    
    def test_apply_size_filter(self):
        with SizeFilterManager(self.temp_file.name, self.config) as manager:
            selected_sizes = ["044", "045"]
            hidden_count = manager.apply_size_filter(selected_sizes)
            self.assertGreater(hidden_count, 0)
            
            for row in range(19, 60):
                cell_value = manager.ws[f'F{row}'].value
                if cell_value and str(cell_value).strip().zfill(3) in selected_sizes:
                    self.assertFalse(manager.ws.row_dimensions[row].hidden)
    
    def test_reset_all_rows(self):
        with SizeFilterManager(self.temp_file.name, self.config) as manager:
            manager.apply_size_filter(["044"])
            manager.reset_all_rows()
            
            for row in range(19, 60):
                self.assertFalse(manager.ws.row_dimensions[row].hidden)
    
    def test_validation_prevents_hiding_outside_range(self):
        with SizeFilterManager(self.temp_file.name, self.config) as manager:
            manager.apply_size_filter(["044"])
            
            for row in range(1, 19):
                self.assertFalse(manager.ws.row_dimensions[row].hidden)
            
            for row in range(60, 70):
                self.assertFalse(manager.ws.row_dimensions[row].hidden)


if __name__ == "__main__":
    unittest.main()

