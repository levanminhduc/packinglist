"""
Module định dạng Excel (màu sắc, font, borders, etc).
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Optional, List
import logging

logger = logging.getLogger(__name__)


class ExcelFormatter:
    """Class định dạng Excel file."""
    
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"File không tồn tại: {file_path}")
        logger.info(f"Khởi tạo ExcelFormatter cho file: {file_path}")
    
    def format_header(
        self,
        sheet_name: Optional[str] = None,
        font_size: int = 12,
        bold: bool = True,
        bg_color: str = "366092",
        font_color: str = "FFFFFF"
    ) -> None:
        """
        Định dạng header row.
        
        Args:
            sheet_name: Tên sheet
            font_size: Kích thước font
            bold: In đậm
            bg_color: Màu nền (hex)
            font_color: Màu chữ (hex)
        """
        try:
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb[sheet_name] if sheet_name else wb.active
            
            header_font = Font(size=font_size, bold=bold, color=font_color)
            header_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            wb.save(self.file_path)
            wb.close()
            logger.info(f"Định dạng header cho sheet '{ws.title}'")
        except Exception as e:
            logger.error(f"Lỗi khi định dạng header: {e}")
            raise
    
    def auto_adjust_column_width(
        self,
        sheet_name: Optional[str] = None,
        min_width: int = 10,
        max_width: int = 50
    ) -> None:
        """
        Tự động điều chỉnh độ rộng cột.
        
        Args:
            sheet_name: Tên sheet
            min_width: Độ rộng tối thiểu
            max_width: Độ rộng tối đa
        """
        try:
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb[sheet_name] if sheet_name else wb.active
            
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                adjusted_width = min(max(max_length + 2, min_width), max_width)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(self.file_path)
            wb.close()
            logger.info(f"Điều chỉnh độ rộng cột cho sheet '{ws.title}'")
        except Exception as e:
            logger.error(f"Lỗi khi điều chỉnh độ rộng cột: {e}")
            raise
    
    def add_borders(
        self,
        sheet_name: Optional[str] = None,
        border_style: str = 'thin'
    ) -> None:
        """
        Thêm viền cho tất cả các ô có dữ liệu.
        
        Args:
            sheet_name: Tên sheet
            border_style: Kiểu viền ('thin', 'medium', 'thick')
        """
        try:
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb[sheet_name] if sheet_name else wb.active
            
            side = Side(style=border_style, color="000000")
            border = Border(left=side, right=side, top=side, bottom=side)
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    if cell.value is not None:
                        cell.border = border
            
            wb.save(self.file_path)
            wb.close()
            logger.info(f"Thêm viền cho sheet '{ws.title}'")
        except Exception as e:
            logger.error(f"Lỗi khi thêm viền: {e}")
            raise
    
    def highlight_cells(
        self,
        sheet_name: Optional[str] = None,
        condition: callable = None,
        bg_color: str = "FFFF00"
    ) -> None:
        """
        Highlight các ô thỏa điều kiện.
        
        Args:
            sheet_name: Tên sheet
            condition: Function kiểm tra điều kiện
            bg_color: Màu nền (hex)
        """
        try:
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb[sheet_name] if sheet_name else wb.active
            
            fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    if cell.value is not None and condition and condition(cell.value):
                        cell.fill = fill
            
            wb.save(self.file_path)
            wb.close()
            logger.info(f"Highlight cells cho sheet '{ws.title}'")
        except Exception as e:
            logger.error(f"Lỗi khi highlight cells: {e}")
            raise
    
    def freeze_panes(
        self,
        sheet_name: Optional[str] = None,
        row: int = 1,
        col: int = 0
    ) -> None:
        """
        Freeze panes (cố định hàng/cột).
        
        Args:
            sheet_name: Tên sheet
            row: Số hàng cần freeze (từ trên xuống)
            col: Số cột cần freeze (từ trái sang)
        """
        try:
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb[sheet_name] if sheet_name else wb.active
            
            freeze_cell = ws.cell(row=row + 1, column=col + 1)
            ws.freeze_panes = freeze_cell.coordinate
            
            wb.save(self.file_path)
            wb.close()
            logger.info(f"Freeze panes tại {freeze_cell.coordinate}")
        except Exception as e:
            logger.error(f"Lỗi khi freeze panes: {e}")
            raise
    
    def apply_number_format(
        self,
        sheet_name: Optional[str] = None,
        columns: List[str] = None,
        format_string: str = "#,##0.00"
    ) -> None:
        """
        Áp dụng định dạng số.
        
        Args:
            sheet_name: Tên sheet
            columns: Danh sách cột cần format (vd: ['A', 'B'])
            format_string: Chuỗi định dạng
        """
        try:
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb[sheet_name] if sheet_name else wb.active
            
            if columns:
                for col_letter in columns:
                    for cell in ws[col_letter]:
                        if cell.row > 1:
                            cell.number_format = format_string
            
            wb.save(self.file_path)
            wb.close()
            logger.info(f"Áp dụng number format cho cột {columns}")
        except Exception as e:
            logger.error(f"Lỗi khi áp dụng number format: {e}")
            raise

