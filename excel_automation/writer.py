"""
Module ghi file Excel với nhiều phương thức khác nhau.
"""

import pandas as pd
import xlsxwriter
import openpyxl
from pathlib import Path
from typing import Optional, List, Dict, Any
import logging

logger = logging.getLogger(__name__)


class ExcelWriter:
    """Class để ghi file Excel với nhiều phương thức."""
    
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.file_path.parent.mkdir(parents=True, exist_ok=True)
        logger.info(f"Khởi tạo ExcelWriter cho file: {file_path}")
    
    def write_dataframe(
        self,
        df: pd.DataFrame,
        sheet_name: str = 'Sheet1',
        index: bool = False,
        header: bool = True,
        mode: str = 'w'
    ) -> None:
        """
        Ghi DataFrame ra Excel bằng pandas.
        
        Args:
            df: DataFrame cần ghi
            sheet_name: Tên sheet
            index: Có ghi index không
            header: Có ghi header không
            mode: 'w' = ghi đè, 'a' = append
        """
        try:
            with pd.ExcelWriter(
                self.file_path,
                engine='openpyxl',
                mode=mode
            ) as writer:
                df.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=index,
                    header=header
                )
            logger.info(f"Ghi thành công {len(df)} dòng vào sheet '{sheet_name}'")
        except Exception as e:
            logger.error(f"Lỗi khi ghi DataFrame: {e}")
            raise
    
    def write_multiple_sheets(
        self,
        data_dict: Dict[str, pd.DataFrame],
        index: bool = False
    ) -> None:
        """
        Ghi nhiều DataFrame vào nhiều sheet.
        
        Args:
            data_dict: Dictionary với key là tên sheet, value là DataFrame
            index: Có ghi index không
        """
        try:
            with pd.ExcelWriter(self.file_path, engine='openpyxl') as writer:
                for sheet_name, df in data_dict.items():
                    df.to_excel(
                        writer,
                        sheet_name=sheet_name,
                        index=index
                    )
            logger.info(f"Ghi thành công {len(data_dict)} sheets")
        except Exception as e:
            logger.error(f"Lỗi khi ghi nhiều sheets: {e}")
            raise
    
    def write_with_xlsxwriter(
        self,
        data: List[List[Any]],
        sheet_name: str = 'Sheet1',
        headers: Optional[List[str]] = None
    ) -> xlsxwriter.Workbook:
        """
        Ghi Excel bằng xlsxwriter - phù hợp cho file lớn và formatting phức tạp.
        
        Args:
            data: Dữ liệu dạng list 2D
            sheet_name: Tên sheet
            headers: Danh sách header
            
        Returns:
            Workbook object
        """
        try:
            workbook = xlsxwriter.Workbook(str(self.file_path))
            worksheet = workbook.add_worksheet(sheet_name)
            
            row_offset = 0
            if headers:
                worksheet.write_row(0, 0, headers)
                row_offset = 1
            
            for row_idx, row_data in enumerate(data):
                worksheet.write_row(row_idx + row_offset, 0, row_data)
            
            logger.info(f"Ghi thành công {len(data)} dòng với xlsxwriter")
            return workbook
        except Exception as e:
            logger.error(f"Lỗi khi ghi với xlsxwriter: {e}")
            raise
    
    def append_dataframe(
        self,
        df: pd.DataFrame,
        sheet_name: str = 'Sheet1'
    ) -> None:
        """
        Thêm DataFrame vào cuối sheet hiện có.
        
        Args:
            df: DataFrame cần thêm
            sheet_name: Tên sheet
        """
        try:
            if not self.file_path.exists():
                self.write_dataframe(df, sheet_name)
                return
            
            wb = openpyxl.load_workbook(self.file_path)
            
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                start_row = ws.max_row + 1
            else:
                ws = wb.create_sheet(sheet_name)
                start_row = 1
                for col_idx, col_name in enumerate(df.columns, 1):
                    ws.cell(row=1, column=col_idx, value=col_name)
                start_row = 2
            
            for row_idx, row_data in enumerate(df.values, start_row):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            wb.save(self.file_path)
            wb.close()
            logger.info(f"Append thành công {len(df)} dòng vào sheet '{sheet_name}'")
        except Exception as e:
            logger.error(f"Lỗi khi append DataFrame: {e}")
            raise
    
    def write_dict_to_excel(
        self,
        data: List[Dict[str, Any]],
        sheet_name: str = 'Sheet1'
    ) -> None:
        """
        Ghi list of dictionaries ra Excel.
        
        Args:
            data: List các dictionary
            sheet_name: Tên sheet
        """
        try:
            df = pd.DataFrame(data)
            self.write_dataframe(df, sheet_name)
            logger.info(f"Ghi thành công {len(data)} records từ dictionary")
        except Exception as e:
            logger.error(f"Lỗi khi ghi dictionary: {e}")
            raise
    
    def create_empty_workbook(
        self,
        sheet_names: List[str] = None
    ) -> None:
        """
        Tạo workbook rỗng với các sheet.
        
        Args:
            sheet_names: Danh sách tên sheet cần tạo
        """
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            
            if sheet_names:
                for sheet_name in sheet_names:
                    wb.create_sheet(sheet_name)
            else:
                wb.create_sheet('Sheet1')
            
            wb.save(self.file_path)
            wb.close()
            logger.info(f"Tạo workbook rỗng với {len(sheet_names or [1])} sheets")
        except Exception as e:
            logger.error(f"Lỗi khi tạo workbook rỗng: {e}")
            raise

