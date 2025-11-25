"""
Module đọc file Excel với nhiều phương thức khác nhau.
"""

import pandas as pd
import openpyxl
from pathlib import Path
from typing import Optional, List, Dict, Any
import logging

logger = logging.getLogger(__name__)


class ExcelReader:
    """Class để đọc file Excel với nhiều phương thức."""
    
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"File không tồn tại: {file_path}")
        logger.info(f"Khởi tạo ExcelReader cho file: {file_path}")
    
    def read_with_pandas(
        self,
        sheet_name: Optional[str] = None,
        header: int = 0,
        usecols: Optional[List[str]] = None
    ) -> pd.DataFrame:
        """
        Đọc Excel bằng pandas - phù hợp cho phân tích dữ liệu.

        Args:
            sheet_name: Tên sheet cần đọc (None = sheet đầu tiên)
            header: Dòng làm header (0 = dòng đầu tiên)
            usecols: Danh sách cột cần đọc

        Returns:
            DataFrame chứa dữ liệu
        """
        try:
            if sheet_name is None:
                sheet_name = 0

            df = pd.read_excel(
                self.file_path,
                sheet_name=sheet_name,
                header=header,
                usecols=usecols,
                engine='openpyxl'
            )

            if isinstance(df, dict):
                df = list(df.values())[0]

            logger.info(f"Đọc thành công {len(df)} dòng từ sheet '{sheet_name}'")
            return df
        except Exception as e:
            logger.error(f"Lỗi khi đọc file: {e}")
            raise
    
    def read_with_openpyxl(
        self, 
        sheet_name: Optional[str] = None,
        read_only: bool = False
    ) -> openpyxl.worksheet.worksheet.Worksheet:
        """
        Đọc Excel bằng openpyxl - phù hợp khi cần giữ formatting.
        
        Args:
            sheet_name: Tên sheet cần đọc
            read_only: Chế độ read-only (tiết kiệm bộ nhớ)
            
        Returns:
            Worksheet object
        """
        try:
            wb = openpyxl.load_workbook(
                self.file_path,
                read_only=read_only,
                data_only=True
            )
            
            if sheet_name:
                ws = wb[sheet_name]
            else:
                ws = wb.active
            
            logger.info(f"Đọc thành công sheet '{ws.title}'")
            return ws
        except Exception as e:
            logger.error(f"Lỗi khi đọc file với openpyxl: {e}")
            raise
    
    def get_sheet_names(self) -> List[str]:
        """Lấy danh sách tên các sheet trong file."""
        try:
            wb = openpyxl.load_workbook(self.file_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            logger.info(f"File có {len(sheet_names)} sheets: {sheet_names}")
            return sheet_names
        except Exception as e:
            logger.error(f"Lỗi khi lấy danh sách sheet: {e}")
            raise
    
    def read_all_sheets(self) -> Dict[str, pd.DataFrame]:
        """
        Đọc tất cả các sheet trong file.
        
        Returns:
            Dictionary với key là tên sheet, value là DataFrame
        """
        try:
            all_sheets = pd.read_excel(
                self.file_path,
                sheet_name=None,
                engine='openpyxl'
            )
            logger.info(f"Đọc thành công {len(all_sheets)} sheets")
            return all_sheets
        except Exception as e:
            logger.error(f"Lỗi khi đọc tất cả sheets: {e}")
            raise
    
    def read_range(
        self, 
        sheet_name: Optional[str] = None,
        start_row: int = 1,
        end_row: Optional[int] = None,
        start_col: int = 1,
        end_col: Optional[int] = None
    ) -> List[List[Any]]:
        """
        Đọc một vùng dữ liệu cụ thể.
        
        Args:
            sheet_name: Tên sheet
            start_row: Dòng bắt đầu (1-indexed)
            end_row: Dòng kết thúc
            start_col: Cột bắt đầu (1-indexed)
            end_col: Cột kết thúc
            
        Returns:
            List 2D chứa dữ liệu
        """
        try:
            ws = self.read_with_openpyxl(sheet_name, read_only=True)
            
            if end_row is None:
                end_row = ws.max_row
            if end_col is None:
                end_col = ws.max_column
            
            data = []
            for row in ws.iter_rows(
                min_row=start_row,
                max_row=end_row,
                min_col=start_col,
                max_col=end_col,
                values_only=True
            ):
                data.append(list(row))
            
            logger.info(f"Đọc vùng dữ liệu: {len(data)} dòng x {len(data[0]) if data else 0} cột")
            return data
        except Exception as e:
            logger.error(f"Lỗi khi đọc vùng dữ liệu: {e}")
            raise
    
    def get_cell_value(
        self, 
        cell: str,
        sheet_name: Optional[str] = None
    ) -> Any:
        """
        Lấy giá trị của một ô cụ thể.
        
        Args:
            cell: Địa chỉ ô (vd: 'A1', 'B5')
            sheet_name: Tên sheet
            
        Returns:
            Giá trị của ô
        """
        try:
            ws = self.read_with_openpyxl(sheet_name, read_only=True)
            value = ws[cell].value
            logger.info(f"Giá trị ô {cell}: {value}")
            return value
        except Exception as e:
            logger.error(f"Lỗi khi đọc ô {cell}: {e}")
            raise

