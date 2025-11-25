import openpyxl
from typing import List, Set, Dict, Optional
from pathlib import Path
import logging

from excel_automation.size_filter_config import SizeFilterConfig
from excel_automation.utils import get_size_sort_key

logger = logging.getLogger(__name__)


class SizeFilterManager:
    
    def __init__(self, file_path: str, config: Optional[SizeFilterConfig] = None):
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"File không tồn tại: {file_path}")
        
        self.config = config or SizeFilterConfig()
        self.wb = None
        self.ws = None
        
        logger.info(f"Khởi tạo SizeFilterManager cho file: {file_path}")
    
    def _load_workbook(self) -> None:
        if self.wb is None:
            self.wb = openpyxl.load_workbook(self.file_path)
            sheet_name = self.config.get_sheet_name()
            
            if sheet_name in self.wb.sheetnames:
                self.ws = self.wb[sheet_name]
            else:
                self.ws = self.wb.active
                logger.warning(f"Sheet '{sheet_name}' không tồn tại, sử dụng sheet active: {self.ws.title}")
    
    def _validate_row_range(self, start_row: int, end_row: int) -> None:
        if start_row < 1:
            raise ValueError("Dòng bắt đầu phải >= 1")
        
        if start_row >= end_row:
            raise ValueError(f"Dòng bắt đầu ({start_row}) phải < dòng kết thúc ({end_row})")
        
        self._load_workbook()
        max_row = self.ws.max_row
        
        if end_row > max_row:
            raise ValueError(
                f"Dòng kết thúc ({end_row}) vượt quá số dòng thực tế trong sheet ({max_row})"
            )
    
    def scan_sizes(self, column: Optional[str] = None, start_row: Optional[int] = None, 
                   end_row: Optional[int] = None) -> List[str]:
        column = column or self.config.get_column()
        start_row = start_row or self.config.get_start_row()
        end_row = end_row or self.config.get_end_row()
        
        self._validate_row_range(start_row, end_row)
        self._load_workbook()
        
        sizes = set()
        
        for row in range(start_row, end_row + 1):
            cell_value = self.ws[f'{column}{row}'].value
            
            if cell_value is not None:
                size_str = str(cell_value).strip()
                
                if size_str.isdigit():
                    size_str = size_str.zfill(3)
                
                if size_str:
                    sizes.add(size_str)

        sorted_sizes = sorted(sizes, key=get_size_sort_key)
        logger.info(f"Quét được {len(sorted_sizes)} size khác nhau trong {column}[{start_row}:{end_row}]")

        return sorted_sizes
    
    def get_size_row_mapping(self, column: Optional[str] = None, start_row: Optional[int] = None,
                            end_row: Optional[int] = None) -> Dict[str, List[int]]:
        column = column or self.config.get_column()
        start_row = start_row or self.config.get_start_row()
        end_row = end_row or self.config.get_end_row()
        
        self._validate_row_range(start_row, end_row)
        self._load_workbook()
        
        size_rows = {}
        
        for row in range(start_row, end_row + 1):
            cell_value = self.ws[f'{column}{row}'].value
            
            if cell_value is not None:
                size_str = str(cell_value).strip()
                if size_str.isdigit():
                    size_str = size_str.zfill(3)
                
                if size_str:
                    if size_str not in size_rows:
                        size_rows[size_str] = []
                    size_rows[size_str].append(row)
        
        return size_rows
    
    def apply_size_filter(self, selected_sizes: List[str], column: Optional[str] = None,
                         start_row: Optional[int] = None, end_row: Optional[int] = None) -> int:
        column = column or self.config.get_column()
        start_row = start_row or self.config.get_start_row()
        end_row = end_row or self.config.get_end_row()
        
        self._validate_row_range(start_row, end_row)
        self._load_workbook()
        
        selected_set = set(selected_sizes)
        hidden_count = 0
        
        for row in range(start_row, end_row + 1):
            cell_value = self.ws[f'{column}{row}'].value
            
            if cell_value is not None:
                size_str = str(cell_value).strip()
                if size_str.isdigit():
                    size_str = size_str.zfill(3)
                
                if size_str not in selected_set:
                    self.ws.row_dimensions[row].hidden = True
                    hidden_count += 1
                else:
                    self.ws.row_dimensions[row].hidden = False
            else:
                self.ws.row_dimensions[row].hidden = True
                hidden_count += 1
        
        logger.info(f"Đã ẩn {hidden_count} dòng, hiển thị {(end_row - start_row + 1) - hidden_count} dòng")
        return hidden_count
    
    def reset_all_rows(self, start_row: Optional[int] = None, end_row: Optional[int] = None) -> None:
        start_row = start_row or self.config.get_start_row()
        end_row = end_row or self.config.get_end_row()
        
        self._validate_row_range(start_row, end_row)
        self._load_workbook()
        
        for row in range(start_row, end_row + 1):
            self.ws.row_dimensions[row].hidden = False
        
        logger.info(f"Đã hiện lại tất cả dòng từ {start_row} đến {end_row}")
    
    def save(self, output_path: Optional[str] = None) -> None:
        if self.wb is None:
            raise RuntimeError("Chưa load workbook")
        
        save_path = output_path or self.file_path
        self.wb.save(save_path)
        logger.info(f"Đã lưu file: {save_path}")
    
    def close(self) -> None:
        if self.wb:
            self.wb.close()
            self.wb = None
            self.ws = None
            logger.info("Đã đóng workbook")
    
    def __enter__(self):
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()

