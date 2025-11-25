from dataclasses import dataclass, field
from typing import List, Dict, Any, Optional
import pandas as pd
from pathlib import Path
import json
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

from excel_automation.validation_rules import ValidationRule, UniqueRule
from excel_automation.writer import ExcelWriter
from excel_automation.formatter import ExcelFormatter

logger = logging.getLogger(__name__)


@dataclass
class ValidationError:
    row_index: int
    column: str
    value: Any
    rule: str
    message: str
    
    def to_dict(self) -> Dict[str, Any]:
        return {
            'row_index': self.row_index,
            'column': self.column,
            'value': self.value,
            'rule': self.rule,
            'message': self.message
        }


@dataclass
class ValidationResult:
    is_valid: bool
    total_rows: int
    error_count: int
    errors: List[ValidationError] = field(default_factory=list)
    
    @property
    def summary(self) -> Dict[str, Any]:
        error_by_column = {}
        for error in self.errors:
            if error.column not in error_by_column:
                error_by_column[error.column] = 0
            error_by_column[error.column] += 1
        
        return {
            'is_valid': self.is_valid,
            'total_rows': self.total_rows,
            'error_count': self.error_count,
            'valid_rows': self.total_rows - len(set(e.row_index for e in self.errors)),
            'errors_by_column': error_by_column
        }
    
    def to_dict(self) -> Dict[str, Any]:
        return {
            'is_valid': self.is_valid,
            'total_rows': self.total_rows,
            'error_count': self.error_count,
            'errors': [e.to_dict() for e in self.errors],
            'summary': self.summary
        }


class DataValidator:
    
    def __init__(self, rules: Optional[Dict[str, List[ValidationRule]]] = None):
        self.rules = rules or {}
        logger.info(f"Khởi tạo DataValidator với {len(self.rules)} cột có rules")
    
    def add_rule(self, column: str, rule: ValidationRule) -> None:
        if column not in self.rules:
            self.rules[column] = []
        self.rules[column].append(rule)
        logger.debug(f"Thêm rule {rule.__class__.__name__} cho cột '{column}'")
    
    def add_rules(self, column: str, rules: List[ValidationRule]) -> None:
        for rule in rules:
            self.add_rule(column, rule)
    
    def validate_dataframe(self, df: pd.DataFrame) -> ValidationResult:
        logger.info(f"Bắt đầu validate DataFrame với {len(df)} dòng")
        
        errors = []
        total_rows = len(df)
        
        for column, rules in self.rules.items():
            if column not in df.columns:
                logger.warning(f"Cột '{column}' không tồn tại trong DataFrame")
                continue
            
            for rule in rules:
                if isinstance(rule, UniqueRule):
                    rule.reset()
            
            for idx, value in enumerate(df[column]):
                row_data = df.iloc[idx].to_dict()
                
                for rule in rules:
                    is_valid, error_message = rule.validate(value, row_data)
                    
                    if not is_valid:
                        error = ValidationError(
                            row_index=idx + 2,
                            column=column,
                            value=value,
                            rule=rule.__class__.__name__,
                            message=error_message
                        )
                        errors.append(error)
                        logger.debug(f"Lỗi tại dòng {idx + 2}, cột '{column}': {error_message}")
        
        is_valid = len(errors) == 0
        result = ValidationResult(
            is_valid=is_valid,
            total_rows=total_rows,
            error_count=len(errors),
            errors=errors
        )
        
        logger.info(f"Hoàn thành validation: {result.error_count} lỗi trong {total_rows} dòng")
        return result
    
    def generate_error_report(
        self, 
        validation_result: ValidationResult, 
        output_path: str
    ) -> None:
        logger.info(f"Tạo báo cáo lỗi tại: {output_path}")
        
        if validation_result.is_valid:
            logger.info("Không có lỗi, không tạo báo cáo")
            return
        
        error_data = []
        for error in validation_result.errors:
            error_data.append({
                'Dòng': error.row_index,
                'Cột': error.column,
                'Giá trị': error.value,
                'Quy tắc vi phạm': error.rule,
                'Thông báo lỗi': error.message
            })
        
        df_errors = pd.DataFrame(error_data)
        
        writer = ExcelWriter(output_path)
        writer.write_dataframe(df_errors, sheet_name='Lỗi Validation')
        
        formatter = ExcelFormatter(output_path)
        formatter.format_header(
            sheet_name='Lỗi Validation',
            bg_color="FF0000",
            font_color="FFFFFF"
        )
        formatter.auto_adjust_column_width(sheet_name='Lỗi Validation')
        formatter.add_borders(sheet_name='Lỗi Validation')
        
        logger.info(f"Đã tạo báo cáo lỗi với {len(error_data)} lỗi")
    
    def highlight_errors_in_excel(
        self, 
        input_path: str,
        validation_result: ValidationResult,
        output_path: str,
        sheet_name: Optional[str] = None
    ) -> None:
        logger.info(f"Highlight lỗi trong file Excel: {input_path}")
        
        if validation_result.is_valid:
            logger.info("Không có lỗi, không highlight")
            return
        
        import shutil
        shutil.copy(input_path, output_path)
        
        wb = load_workbook(output_path)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        error_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        error_font = Font(color="FF0000", bold=True)
        
        error_cells = {}
        for error in validation_result.errors:
            key = (error.row_index, error.column)
            if key not in error_cells:
                error_cells[key] = []
            error_cells[key].append(error.message)
        
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        column_indices = {col: idx + 1 for idx, col in enumerate(header_row)}
        
        for (row_idx, column), messages in error_cells.items():
            if column in column_indices:
                col_idx = column_indices[column]
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = error_fill
                cell.font = error_font
                
                comment_text = "\n".join(messages)
                from openpyxl.comments import Comment
                cell.comment = Comment(comment_text, "Validator")
        
        wb.save(output_path)
        wb.close()
        
        logger.info(f"Đã highlight {len(error_cells)} ô có lỗi")
    
    @classmethod
    def from_json(cls, json_path: str) -> 'DataValidator':
        logger.info(f"Load validation rules từ JSON: {json_path}")
        
        with open(json_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
        
        validator = cls()
        
        from excel_automation.validation_rules import (
            RequiredRule, TypeRule, RangeRule, RegexRule, 
            LengthRule, DateRule, UniqueRule, InSetRule
        )
        
        rule_classes = {
            'required': RequiredRule,
            'type': TypeRule,
            'range': RangeRule,
            'regex': RegexRule,
            'length': LengthRule,
            'date': DateRule,
            'unique': UniqueRule,
            'in_set': InSetRule
        }
        
        for column, rules_config in config.items():
            for rule_config in rules_config:
                rule_type = rule_config.get('type')
                rule_class = rule_classes.get(rule_type)
                
                if not rule_class:
                    logger.warning(f"Không tìm thấy rule type: {rule_type}")
                    continue
                
                params = rule_config.get('params', {})
                error_message = rule_config.get('error_message')
                
                if rule_type == 'required':
                    rule = RequiredRule(column, error_message)
                elif rule_type == 'type':
                    type_map = {'int': int, 'float': float, 'str': str}
                    expected_type = type_map.get(params.get('expected_type', 'str'))
                    rule = TypeRule(column, expected_type, error_message)
                elif rule_type == 'range':
                    rule = RangeRule(
                        column, 
                        params.get('min_value'), 
                        params.get('max_value'),
                        error_message
                    )
                elif rule_type == 'regex':
                    rule = RegexRule(column, params.get('pattern'), error_message)
                elif rule_type == 'length':
                    rule = LengthRule(
                        column,
                        params.get('min_length'),
                        params.get('max_length'),
                        error_message
                    )
                elif rule_type == 'date':
                    rule = DateRule(
                        column,
                        params.get('date_format', '%Y-%m-%d'),
                        error_message
                    )
                elif rule_type == 'unique':
                    rule = UniqueRule(column, error_message)
                elif rule_type == 'in_set':
                    rule = InSetRule(
                        column,
                        params.get('allowed_values', []),
                        params.get('case_sensitive', True),
                        error_message
                    )
                else:
                    continue
                
                validator.add_rule(column, rule)
        
        logger.info(f"Đã load {sum(len(rules) for rules in validator.rules.values())} rules")
        return validator

